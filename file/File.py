import os, re, io, base64, csv, torch, shutil, requests, chardet, pathlib
import openpyxl, folder_paths, node_helpers
import numpy as np
from PIL import Image, ImageOps, ImageSequence, ImageDraw, ImageFont
from pathlib import Path
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from io import BytesIO



#======全能加载图像
class GenericImageLoader:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image_input": ("STRING", {"default": ""}),
            }
        }
    RETURN_TYPES = ("IMAGE", "MASK")
    FUNCTION = "load_image"
    OUTPUT_NODE = False
    CATEGORY = "🇨🇳BOZO/PIC"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def load_image(self, image_input):
        path = image_input.strip()
        for ctrl in ['\u202a', '\u202b', '\u202c', '\u202d', '\u202e']:
            while path.startswith(ctrl):
                path = path.lstrip(ctrl)

        source = None
        lower = path.lower()
        if lower.startswith('http://') or lower.startswith('https://'):
            source = 'network'
        elif os.path.isfile(path):
            source = 'local'
        else:
            source = 'base64'
        if source == 'local':
            img = Image.open(path)
        elif source == 'network':
            resp = requests.get(path)
            resp.raise_for_status()
            img = Image.open(io.BytesIO(resp.content))
        else:  # 'base64'
            if ',' in path and path.startswith('data:'):
                _, data = path.split(',', 1)
            else:
                data = path
            decoded = base64.b64decode(data)
            img = Image.open(io.BytesIO(decoded))

        img = img.convert('RGBA')
        has_alpha = img.mode == 'RGBA'
        if has_alpha:
            alpha_channel = np.array(img)[:, :, 3]
            mask = (alpha_channel > 0).astype(np.float32)
            mask_tensor = torch.from_numpy(mask).unsqueeze(0).unsqueeze(0) 
        else:
            mask_tensor = torch.zeros((1, 1, img.size[1], img.size[0]), dtype=torch.float32) 
        np_img = np.array(img).astype(np.float32) / 255.0
        img_tensor = torch.from_numpy(np_img).unsqueeze(0)

        return img_tensor, mask_tensor 


#======加载重置图像
class LoadAndAdjustImage:
    @classmethod
    def INPUT_TYPES(s):
        input_dir = folder_paths.get_input_directory()
        files = [f.name for f in Path(input_dir).iterdir() if f.is_file()]
        return {
            "required": {
                "image": (sorted(files), {"image_upload": True}),
                "max_dimension": ("INT", {"default": 0, "min": 0, "max": 4096, "step": 8}),
                "size_option": (["No Change", "Custom", "Million Pixels", "Small", "Medium", "Large", "Large2", "Large3", "Gigantic", "832×480", "480×832",
                    "480P-H(vid 4:3)", "480P-V(vid 3:4)", "720P-H(vid 16:9)", "720P-V(vid 9:16)", 
                    "1080P-H(vid 16:9)", "1080P-V(vid 9:16)"], 
                                {"default": "No Change"})
            }
        }

    RETURN_TYPES = ("IMAGE", "MASK", "STRING")
    RETURN_NAMES = ("image", "mask", "info")
    FUNCTION = "load_image"
    CATEGORY = "🇨🇳BOZO"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def load_image(self, image, max_dimension, size_option):
        image_path = folder_paths.get_annotated_filepath(image)
        img = Image.open(image_path)
        W, H = img.size
        aspect_ratio = W / H

        def get_target_size():
            if size_option == "No Change":
                # No resizing or cropping, just return the original size
                return W, H
            elif size_option == "Million Pixels":
                return self._resize_to_million_pixels(W, H)
            elif size_option == "Custom":
                ratio = min(max_dimension / W, max_dimension / H)
                return round(W * ratio), round(H * ratio)
            
            size_options = {
                "Small": (
                    (768, 512) if aspect_ratio >= 1.23 else
                    (512, 768) if aspect_ratio <= 0.82 else
                    (768, 768)
                ),
                "Medium": (
                    (1216, 832) if aspect_ratio >= 1.23 else
                    (832, 1216) if aspect_ratio <= 0.82 else
                    (1216, 1216)
                ),
                "Large": (
                    (1600, 1120) if aspect_ratio >= 1.23 else
                    (1120, 1600) if aspect_ratio <= 0.82 else
                    (1600, 1600)
                ),
                # 添加新的Large尺寸选项
                "Large2": (
                    (2048, 1408) if aspect_ratio >= 1.23 else
                    (1408, 2048) if aspect_ratio <= 0.82 else
                    (2048, 2048)
                ),
                 "Large3": (
                    (3072, 2112) if aspect_ratio >= 1.23 else
                    (2112, 3072) if aspect_ratio <= 0.82 else
                    (3072, 3072)
                ),
                 "Gigantic": (
                    (4096, 2816) if aspect_ratio >= 1.23 else
                    (2816, 4096) if aspect_ratio <= 0.82 else
                    (4096, 4096)
                ),
                "Million Pixels": self._resize_to_million_pixels(W, H),  # Million Pixels option
                "480P-H(vid 4:3)": (640, 480),  # 480P-H, 640x480
                "480P-V(vid 3:4)": (480, 640),  # 480P-V, 480x640
                "720P-H(vid 16:9)": (1280, 720),  # 720P-H, 1280x720
                "720P-V(vid 9:16)": (720, 1280),  # 720P-V, 720x1280
                "832×480": (832, 480),  # 832x480
                "480×832": (480, 832),  # 480x832
                # 添加新的1080P选项
                "1080P-H(vid 16:9)": (1920, 1080),
                "1080P-V(vid 9:16)": (1080, 1920)
            }
            return size_options[size_option]
        
        target_width, target_height = get_target_size()
        output_images = []
        output_masks = []

        for frame in ImageSequence.Iterator(img):
            frame = ImageOps.exif_transpose(frame)
            if frame.mode == 'P':
                frame = frame.convert("RGBA")
            elif 'A' in frame.getbands():
                frame = frame.convert("RGBA")
            
            if size_option == "No Change":
                # No resizing, just use the original frame
                image_frame = frame.convert("RGB")
            else:
                if size_option == "Custom" or size_option == "Million Pixels":
                    ratio = min(target_width / W, target_height / H)
                    adjusted_width = round(W * ratio)
                    adjusted_height = round(H * ratio)
                    image_frame = frame.convert("RGB").resize((adjusted_width, adjusted_height), Image.Resampling.BILINEAR)
                else:
                    image_frame = frame.convert("RGB")
                    image_frame = ImageOps.fit(image_frame, (target_width, target_height), method=Image.Resampling.BILINEAR, centering=(0.5, 0.5))

            image_array = np.array(image_frame).astype(np.float32) / 255.0
            output_images.append(torch.from_numpy(image_array)[None,])

            # Process the mask if available
            if 'A' in frame.getbands():
                mask_frame = frame.getchannel('A')
                if size_option == "Custom" or size_option == "Million Pixels":
                    mask_frame = mask_frame.resize((adjusted_width, adjusted_height), Image.Resampling.BILINEAR)
                else:
                    mask_frame = ImageOps.fit(mask_frame, (target_width, target_height), method=Image.Resampling.BILINEAR, centering=(0.5, 0.5))
                mask = np.array(mask_frame).astype(np.float32) / 255.0
                mask = 1. - mask
            else:
                if size_option == "Custom" or size_option == "Million Pixels":
                    mask = np.zeros((adjusted_height, adjusted_width), dtype=np.float32)
                else:
                    mask = np.zeros((target_height, target_width), dtype=np.float32)
            output_masks.append(torch.from_numpy(mask).unsqueeze(0))
        
        output_image = torch.cat(output_images, dim=0) if len(output_images) > 1 else output_images[0]
        output_mask = torch.cat(output_masks, dim=0) if len(output_masks) > 1 else output_masks[0]
        info = f"Image Path: {image_path}\nOriginal Size: {W}x{H}\nAdjusted Size: {target_width}x{target_height}"
        return (output_image, output_mask, info)

    @classmethod
    def VALIDATE_INPUTS(s, image):
        if not folder_paths.exists_annotated_filepath(image):
            return f"Invalid image file: {image}"
        return True
    def _resize_to_million_pixels(self, W, H):
        aspect_ratio = W / H
        target_area = 1000000  # 1 million pixels
        if aspect_ratio > 1:  # Landscape
            width = int(np.sqrt(target_area * aspect_ratio))
            height = int(target_area / width)
        else:  # Portrait
            height = int(np.sqrt(target_area / aspect_ratio))
            width = int(target_area / height)
        width = (width + 7) // 8 * 8
        height = (height + 7) // 8 * 8
        return width, height


#======重置图像
class ImageAdjuster:
    @classmethod
    def INPUT_TYPES(s):
        return {
            "required": {
                "image": ("IMAGE",),
                "mask": ("MASK",),  # 添加遮罩输入
                "max_dimension": ("INT", {"default": 0, "min": 0, "max": 4096, "step": 8}),
                "size_option": ([
                    "Custom", "Million Pixels", "Small", "Medium", "Large", "Large2", "Large3", "Gigantic", "832×480", "480×832",
                    "480P-H(vid 4:3)", "480P-V(vid 3:4)", "720P-H(vid 16:9)", "720P-V(vid 9:16)", 
                    "1080P-H(vid 16:9)", "1080P-V(vid 9:16)"], 
                    {"default": "Million Pixels"}
                )
            }
        }

    RETURN_TYPES = ("IMAGE", "MASK", "INT", "INT")
    RETURN_NAMES = ("image", "mask", "width", "height")
    FUNCTION = "process_image"
    CATEGORY = "🇨🇳BOZO"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def process_image(self, image, mask, max_dimension=4096, size_option="Custom"):
        batch_size = image.shape[0]
        processed_images = []
        processed_masks = []
        widths = []
        heights = []

        for i in range(batch_size):
            current_image = image[i]
            current_mask = mask[i]
            
            input_pil_image = Image.fromarray((current_image.numpy() * 255).astype(np.uint8))
            input_pil_mask = Image.fromarray((current_mask.numpy() * 255).astype(np.uint8))
            
            W, H = input_pil_image.size

            processed_image_pil = input_pil_image.copy()
            processed_image_pil = ImageOps.exif_transpose(processed_image_pil)

            processed_mask_pil = input_pil_mask.copy()
            processed_mask_pil = ImageOps.exif_transpose(processed_mask_pil)

            if processed_image_pil.mode == 'P':
                processed_image_pil = processed_image_pil.convert("RGBA")
            elif 'A' in processed_image_pil.getbands():
                processed_image_pil = processed_image_pil.convert("RGBA")

            if processed_mask_pil.mode != "L":
                processed_mask_pil = processed_mask_pil.convert("L")

            if size_option != "Custom":
                aspect_ratio = W / H

                size_options = {
                    "Small": (
                        (768, 512) if aspect_ratio >= 1.23 else
                        (512, 768) if aspect_ratio <= 0.82 else
                        (768, 768)
                    ),
                    "Medium": (
                        (1216, 832) if aspect_ratio >= 1.23 else
                        (832, 1216) if aspect_ratio <= 0.82 else
                        (1216, 1216)
                    ),
                    "Large": (
                        (1600, 1120) if aspect_ratio >= 1.23 else
                        (1120, 1600) if aspect_ratio <= 0.82 else
                        (1600, 1600)
                    ),
                    # 添加新的Large尺寸选项
                    "Large2": (
                        (2048, 1408) if aspect_ratio >= 1.23 else
                        (1408, 2048) if aspect_ratio <= 0.82 else
                        (2048, 2048)
                    ),
                    "Large3": (
                        (3072, 2112) if aspect_ratio >= 1.23 else
                        (2112, 3072) if aspect_ratio <= 0.82 else
                        (3072, 3072)
                    ),
                    "Gigantic": (
                        (4096, 2816) if aspect_ratio >= 1.23 else
                        (2816, 4096) if aspect_ratio <= 0.82 else
                        (4096, 4096)
                    ),
                    "Million Pixels": self._resize_to_million_pixels(W, H),
                    "480P-H(vid 4:3)": (640, 480),
                    "480P-V(vid 3:4)": (480, 640),
                    "720P-H(vid 16:9)": (1280, 720),
                    "720P-V(vid 9:16)": (720, 1280),
                    # 添加新的1080P选项
                    "1080P-H(vid 16:9)": (1920, 1080),
                    "1080P-V(vid 9:16)": (1080, 1920),
                    "832×480": (832, 480),
                    "480×832": (480, 832)
                }

                target_width, target_height = size_options[size_option]
                processed_image_pil = processed_image_pil.convert("RGB")
                processed_image_pil = ImageOps.fit(processed_image_pil, (target_width, target_height), method=Image.Resampling.BILINEAR, centering=(0.5, 0.5))
                
                processed_mask_pil = ImageOps.fit(processed_mask_pil, (target_width, target_height), method=Image.Resampling.BILINEAR, centering=(0.5, 0.5))
            else:
                ratio = min(max_dimension / W, max_dimension / H)
                adjusted_width = round(W * ratio)
                adjusted_height = round(H * ratio)

                processed_image_pil = processed_image_pil.convert("RGB")
                processed_image_pil = processed_image_pil.resize((adjusted_width, adjusted_height), Image.Resampling.BILINEAR)
                
                processed_mask_pil = processed_mask_pil.resize((adjusted_width, adjusted_height), Image.Resampling.BILINEAR)

            processed_image = np.array(processed_image_pil).astype(np.float32) / 255.0
            processed_image = torch.from_numpy(processed_image)
            processed_images.append(processed_image)

            processed_mask = np.array(processed_mask_pil).astype(np.float32) / 255.0
            processed_mask = torch.from_numpy(processed_mask)
            processed_masks.append(processed_mask)

            if size_option != "Custom":
                widths.append(target_width)
                heights.append(target_height)
            else:
                widths.append(adjusted_width)
                heights.append(adjusted_height)

        output_image = torch.stack(processed_images)
        output_mask = torch.stack(processed_masks)
        
        if all(w == widths[0] for w in widths) and all(h == heights[0] for h in heights):
            return (output_image, output_mask, widths[0], heights[0])
        else:
            return (output_image, output_mask, widths[0], heights[0])

    def _resize_to_million_pixels(self, W, H):
        aspect_ratio = W / H
        target_area = 1000000
        
        if aspect_ratio > 1:
            width = int(np.sqrt(target_area * aspect_ratio))
            height = int(target_area / width)
        else:
            height = int(np.sqrt(target_area / aspect_ratio))
            width = int(target_area / height)

        width = (width + 7) // 8 * 8
        height = (height + 7) // 8 * 8
        
        return width, height


#======裁剪图像
class CustomCrop:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image": ("IMAGE",),
                "mask": ("MASK",),  # 新增遮罩输入
                "width": ("INT", {"default": 768, "min": 0, "max": 4096, "step": 8}),
                "height": ("INT", {"default": 768, "min": 0, "max": 4096, "step": 8}),
            }
        }

    RETURN_TYPES = ("IMAGE", "MASK", "INT", "INT")  # 新增遮罩输出
    RETURN_NAMES = ("image", "mask", "width", "height")
    FUNCTION = "process_image"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def process_image(self, image, mask, width=768, height=768):
        input_image = Image.fromarray((image.squeeze(0).numpy() * 255).astype(np.uint8))
        input_mask = Image.fromarray((mask.squeeze(0).numpy() * 255).astype(np.uint8))  # 转换遮罩为PIL图像
        
        W, H = input_image.size

        processed_images = []
        processed_masks = []  # 新增遮罩处理列表

        for frame in [input_image]:
            frame = ImageOps.exif_transpose(frame)

            if frame.mode == 'P':
                frame = frame.convert("RGBA")
            elif 'A' in frame.getbands():
                frame = frame.convert("RGBA")

            processed_image = frame.convert("RGB")
            processed_image = ImageOps.fit(processed_image, (width, height), method=Image.Resampling.BILINEAR, centering=(0.5, 0.5))

            processed_image = np.array(processed_image).astype(np.float32) / 255.0
            processed_image = torch.from_numpy(processed_image)[None,]
            processed_images.append(processed_image)

        # 处理遮罩
        input_mask = ImageOps.exif_transpose(input_mask)
        processed_mask = input_mask.convert("L")  # 确保遮罩是灰度图像
        processed_mask = ImageOps.fit(processed_mask, (width, height), method=Image.Resampling.BILINEAR, centering=(0.5, 0.5))
        processed_mask = np.array(processed_mask).astype(np.float32) / 255.0
        processed_mask = torch.from_numpy(processed_mask)[None,]
        processed_masks.append(processed_mask)

        output_image = torch.cat(processed_images, dim=0) if len(processed_images) > 1 else processed_images[0]
        output_mask = torch.cat(processed_masks, dim=0) if len(processed_masks) > 1 else processed_masks[0]

        return (output_image, output_mask, width, height)


#======保存图像
class SaveImagEX:
    def __init__(self):
        self.output_dir = folder_paths.get_output_directory()

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image": ("IMAGE",),
                "save_path": ("STRING", {"default": "./output"}),
                "image_name": ("STRING", {"default": "A"}),
                "image_format": (["PNG", "JPG", "WEBP"], {"default": "JPG"}),
                "quality": ("INT", {"default": 100, "min": 1, "max": 100, "step": 1})
            },
        }

    RETURN_TYPES = ("IMAGE",)
    RETURN_NAMES = ("image",)
    FUNCTION = "save_image"
    OUTPUT_NODE = True
    CATEGORY = "🇨🇳BOZO"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def save_image(self, image, save_path, image_name, image_format, quality):
        if not isinstance(image, torch.Tensor):
            raise ValueError("Invalid image tensor format")
        if save_path == "./output":
            save_path = self.output_dir
        elif not os.path.isabs(save_path):
            save_path = os.path.join(self.output_dir, save_path)
        os.makedirs(save_path, exist_ok=True)
        
        # 移除可能存在的扩展名，然后添加用户选择的格式对应的扩展名
        base_name = os.path.splitext(image_name)[0]
        
        batch_size = image.shape[0]
        channel_to_mode = {1: "L", 3: "RGB", 4: "RGBA"}

        for i in range(batch_size):
            if image_format == "PNG":
                # 优化文件命名规则，确保所有图片都按照 {image_name}_{序号} 的格式命名
                filename = f"{base_name}_{i+1:05d}.png"
                save_format = "PNG"
                save_params = {"compress_level": 0}
            elif image_format == "JPG":
                filename = f"{base_name}_{i+1:05d}.jpg"
                save_format = "JPEG"
                save_params = {"quality": quality}
            else:  # WEBP
                filename = f"{base_name}_{i+1:05d}.webp"
                save_format = "WEBP"
                save_params = {"quality": quality}
            
            full_path = os.path.join(save_path, filename)
            single_image = image[i].cpu().numpy()
            single_image = (single_image * 255).astype('uint8')
            channels = single_image.shape[2]
            if channels not in channel_to_mode:
                raise ValueError(f"Unsupported channel number: {channels}")
            mode = channel_to_mode[channels]
            if channels == 1:
                single_image = single_image[:, :, 0]
            pil_image = Image.fromarray(single_image, mode)
            
            if image_format == "JPG":
                pil_image = pil_image.convert("RGB")
            elif image_format == "WEBP":
                pil_image = pil_image.convert("RGB") if mode != "RGBA" else pil_image
            
            pil_image.save(full_path, format=save_format, **save_params)
        return (image,)


#======文件操作
class FileCopyCutNode:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "source_path": ("STRING", {"default": "", "multiline": False}),
                "destination_path": ("STRING", {"default": "", "multiline": False}),
                "operation": (["copy", "cut"], {"default": "copy"}),
            },
              # 可以进一步简化成一行
        }

    RETURN_TYPES = ("STRING",)  # 返回操作结果字符串
    RETURN_NAMES = ("result",)
    FUNCTION = "copy_cut_file"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def copy_cut_file(self, source_path, destination_path, operation):
        result = "执行失败"
        try:
            if not os.path.isfile(source_path):
                raise FileNotFoundError(f"源文件未找到: {source_path}")

            os.makedirs(os.path.dirname(destination_path), exist_ok=True)

            if operation == "copy":
                shutil.copy2(source_path, destination_path)
                result = "执行成功: 文件已复制"
            elif operation == "cut":
                shutil.move(source_path, destination_path)
                result = "执行成功: 文件已剪切"
            else:
                raise ValueError("操作类型无效，仅支持 'copy' 或 'cut'。")
        except FileNotFoundError as e:
            result = f"执行失败: {str(e)}"
        except ValueError as e:
            result = f"执行失败: {str(e)}"
        except Exception as e:
            result = f"执行失败: {str(e)}"

        return (result,)


#======替换文件名
class FileNameReplacer:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "file_path": ("STRING", {"default": "path/to/your/file.jpg"}),
                "new_file_name": ("STRING", {"default": ""}),
            },
             
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "replace_file_name"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def replace_file_name(self, file_path, new_file_name):
        dir_name = os.path.dirname(file_path)
        _, file_ext = os.path.splitext(file_path)
        new_file_name = self.sanitize_file_name(new_file_name)
        new_file_path = os.path.join(dir_name, new_file_name + file_ext)
        return (new_file_path,)
    def sanitize_file_name(self, file_name):
        invalid_chars = r'[\/:*?"<>|]'
        return re.sub(invalid_chars, '_', file_name)


#======文本写入TXT
class WriteToTxtFile:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text_content": ("STRING", {"default": "", "multiline": True}),
                "file_path": ("STRING", {"default": "path/to/your/file.txt"}),
            },
             
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "write_to_txt"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def write_to_txt(self, text_content, file_path):
        try:
            dir_path = os.path.dirname(file_path)
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
            file_exists = os.path.exists(file_path)
            mode = 'a' if file_exists else 'w'
            
            with open(file_path, mode, encoding='utf-8') as f:
                if file_exists:
                    f.write('\n\n') 
                f.write(text_content)
            return ("Write successful: " + text_content,)
        except Exception as e:
            return (f"Error: {str(e)}",)


#======清理文件
class FileDeleteNode:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "items_to_delete": ("STRING", {"default": "33.png\ncs1/01.png\ncs1", "multiline": True}),
            },
             
        }
    
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("result",)
    FUNCTION = "delete_files"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def delete_files(self, items_to_delete):
        result = "执行成功: 所有指定项已从output目录删除"
        error_messages = []
        base_output_dir = Path.cwd() / COMFYUI_OUTPUT_DIR
        items = items_to_delete.strip().split('\n')

        for item in items:
            item = item.strip()
            if not item:
                continue
            if item == "[DeleteAll]":
                try:
                    for file_or_dir in base_output_dir.glob('*'):
                        if file_or_dir.is_file() or file_or_dir.is_symlink():
                            file_or_dir.unlink()
                        elif file_or_dir.is_dir():
                            shutil.rmtree(file_or_dir)
                    continue
                except Exception as e:
                    error_messages.append(f"从output目录删除全部失败: {str(e)}")
                    continue
            target_path = base_output_dir / item
            try:
                target_path.relative_to(base_output_dir)
            except ValueError:
                error_messages.append(f"{item} 不在output目录范围内，无法删除")
                continue
            if not target_path.exists():
                error_messages.append(f"在output目录下找不到 {item}")
                continue
            try:
                if target_path.is_file() or target_path.is_symlink():
                    target_path.unlink()
                elif target_path.is_dir():
                    shutil.rmtree(target_path)
            except Exception as e:
                error_messages.append(f"从output目录删除 {item} 失败: {str(e)}")
        if error_messages:
            result = "部分执行失败:\n" + "\n".join(error_messages)
        return (result,)


#======文件路径和后缀统计
class FileListAndSuffix:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "folder_path": ("STRING",),
                "file_extension": (["jpg", "png", "jpg&png", "txt", "csv", "all"], {"default": "jpg"}),
            },
             
        }

    RETURN_TYPES = ("STRING", "INT", "LIST")
    FUNCTION = "file_list_and_suffix"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def file_list_and_suffix(self, folder_path, file_extension):
        try:
            if not os.path.isdir(folder_path):
                return ("", 0, [])

            if file_extension == "all":
                file_paths = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
            elif file_extension == "jpg&png":
                file_paths = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.png'))]
            else:
                file_paths = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith('.' + file_extension)]

            return ('\n'.join(file_paths), len(file_paths), file_paths)
        except Exception as e:
            return ("", 0, [])


#======文字图像
class TextToImage:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {"default": "BOZOYAN", "multiline": True}),
                "font": ("STRING", {"default": "Light.otf"}),
                "max_width": ("INT", {"default": 300, "min": 1, "max": 2048, "step": 1}),
                "font_properties": ("STRING", {"default": "#FFFFFF,1,1,C,1", "multiline": False}),
                "font_stroke": ("STRING", {"default": "#000000,2,1", "multiline": False}),
                "font_background": ("STRING", {"default": "#333333,5,10,1", "multiline": False})
            },
        }
    RETURN_TYPES = ("IMAGE",)
    FUNCTION = "generate_text_image"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def generate_text_image(self, text, font, max_width, font_properties, font_stroke, font_background):
        fonts_dir = os.path.join(os.path.dirname(os.path.realpath(__file__)), "fonts")
        # 判断font是绝对路径还是文件名
        if os.path.isabs(font):
            font_path = font
        else:
            font_path = os.path.join(fonts_dir, font)
        try:
            font_obj = ImageFont.truetype(font_path, 1)
        except Exception as e:
            return None

        draw = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
        lines = text.split("\n")
        max_text_width = 0
        text_height = 0
        for line in lines:
            left, top, right, bottom = draw.textbbox((0, 0), line, font=font_obj)
            line_width = right - left
            line_height = bottom - top
            max_text_width = max(max_text_width, line_width)
            text_height += line_height

        if max_text_width == 0:
            max_text_width = 1

        ratio = max_width / max_text_width
        new_font_size = int(1 * ratio)
        font_obj = ImageFont.truetype(font_path, new_font_size)

        font_color = "#FFFFFF"
        letter_spacing_factor = 1.0
        line_spacing_factor = 1.0
        alignment = "C"
        opacity = 1.0
        stroke_color = "#000000"
        stroke_width = 0.0
        stroke_opacity = 1.0
        background_color = "#333333"
        expand_width = 5
        corner_radius = 10
        background_opacity = 0.9

        try:
            if font_properties.strip() == "":
                font_color = "#FFFFFF"
                letter_spacing_factor = 1.0
                line_spacing_factor = 1.0
                alignment = "C"
                opacity = 1.0
            else:
                props = font_properties.split(',')
                if len(props) >= 5:
                    font_color = props[0].strip()
                    letter_spacing_factor = float(props[1]) if props[1] else 1.0
                    line_spacing_factor = float(props[2]) if props[2] else 1.0
                    alignment = props[3].strip().upper()
                    opacity = float(props[4]) if props[4] else 1.0
                else:
                    font_color = "#FFFFFF"
                    letter_spacing_factor = 1.0
                    line_spacing_factor = 1.0
                    alignment = "C"
                    opacity = 1.0
        except:
            font_color = "#FFFFFF"
            letter_spacing_factor = 1.0
            line_spacing_factor = 1.0
            alignment = "C"
            opacity = 1.0

        try:
            if font_stroke.strip() == "":
                stroke_width = 0.0
            else:
                stroke_props = font_stroke.split(',')
                if len(stroke_props) >= 3:
                    stroke_color = stroke_props[0].strip()
                    stroke_width = float(stroke_props[1]) if stroke_props[1] else 1.0
                    stroke_opacity = float(stroke_props[2]) if stroke_props[2] else 1.0
                else:
                    stroke_width = 0.0
        except:
            stroke_width = 0.0

        try:
            if font_background.strip() == "":
                background_color = None
            else:
                bg_props = font_background.split(',')
                if len(bg_props) >= 4:
                    background_color = bg_props[0].strip()
                    expand_width = int(bg_props[1]) if bg_props[1] else 5
                    corner_radius = int(bg_props[2]) if bg_props[2] else 10
                    background_opacity = float(bg_props[3]) if bg_props[3] else 0.9
                else:
                    background_color = None
        except:
            background_color = None

        actual_max_width = 0
        for line in lines:
            line_width = 0
            for char in line:
                char_width = draw.textbbox((0, 0), char, font=font_obj)[2]
                line_width += char_width + (font_obj.size * (letter_spacing_factor - 1))
            actual_max_width = max(actual_max_width, line_width)

        if actual_max_width > max_width:
            ratio = max_width / actual_max_width
            new_font_size = int(new_font_size * ratio)
            font_obj = ImageFont.truetype(font_path, new_font_size)

        font_ascent, font_descent = font_obj.getmetrics()
        line_height = font_ascent + font_descent

        if len(lines) > 1:
            text_height = line_height * (len(lines) - 1) * line_spacing_factor + line_height
        else:
            text_height = line_height

        image_width = max_width
        image_height = int(text_height + new_font_size * 0.2)
        image = Image.new('RGBA', (image_width, image_height), (0, 0, 0, 0))
        draw = ImageDraw.Draw(image)

        if background_color is not None:
            try:
                background_color_tuple = (
                    int(background_color[1:3], 16),
                    int(background_color[3:5], 16),
                    int(background_color[5:7], 16),
                    int(background_opacity * 255)
                )
                draw.rounded_rectangle(
                    [0, 0, image_width, image_height],
                    fill=background_color_tuple,
                    radius=corner_radius
                )
            except:
                pass

        y_text = new_font_size * 0.1
        try:
            font_color_tuple = (
                int(font_color[1:3], 16),
                int(font_color[3:5], 16),
                int(font_color[5:7], 16),
                int(opacity * 255)
            )
            stroke_color_tuple = (
                int(stroke_color[1:3], 16),
                int(stroke_color[3:5], 16),
                int(stroke_color[5:7], 16),
                int(stroke_opacity * 255)
            )
        except:
            font_color_tuple = (255, 255, 255, 255)
            stroke_color_tuple = (0, 0, 0, 255)

        for i, line in enumerate(lines):
            line_width = 0
            for char in line:
                char_width = draw.textbbox((0, 0), char, font=font_obj)[2]
                line_width += char_width + (font_obj.size * (letter_spacing_factor - 1))
            line_width = max(line_width, 1)

            if alignment == "L":
                x = 0
            elif alignment == "R":
                x = max_width - line_width
            else:
                x = (max_width - line_width) / 2

            if stroke_width > 0:
                for sx in range(-int(stroke_width), int(stroke_width) + 1):
                    for sy in range(-int(stroke_width), int(stroke_width) + 1):
                        if sx == 0 and sy == 0:
                            continue
                        char_x = x + sx
                        char_y = y_text + sy
                        for char in line:
                            char_width = draw.textbbox((0, 0), char, font=font_obj)[2]
                            draw.text((char_x, char_y), char, font=font_obj, fill=stroke_color_tuple)
                            char_x += char_width + (font_obj.size * (letter_spacing_factor - 1))

            char_x = x
            for char in line:
                char_width = draw.textbbox((0, 0), char, font=font_obj)[2]
                draw.text((char_x, y_text), char, font=font_obj, fill=font_color_tuple)
                char_x += char_width + (font_obj.size * (letter_spacing_factor - 1))

            if i < len(lines) - 1:
                y_text += line_height * line_spacing_factor
            else:
                y_text += line_height

        image_data = np.array(image)
        alpha_channel = image_data[:, :, 3]
        non_zero_indices = np.where(alpha_channel > 0)
        if len(non_zero_indices[0]) > 0:
            min_y = np.min(non_zero_indices[0])
            max_y = np.max(non_zero_indices[0])
            min_x = np.min(non_zero_indices[1])
            max_x = np.max(non_zero_indices[1])
            image = image.crop((min_x, min_y, max_x + 1, max_y + 1))
        else:
            pass

        text_content_width = max_x - min_x + 1 if len(non_zero_indices[0]) > 0 else max_width

        image_width, image_height = image.size
        if text_content_width < max_width:
            new_image = Image.new('RGBA', (max_width, image_height), (0, 0, 0, 0))
            new_draw = ImageDraw.Draw(new_image)
            x_offset = (max_width - text_content_width) // 2
            new_image.paste(image, (x_offset, 0))
            image = new_image

        image_width, image_height = image.size
        if image_width > max_width:
            height_ratio = image_height / image_width
            image = image.resize((max_width, int(max_width * height_ratio)))

        image_np = np.array(image).astype(np.float32) / 255.0
        image_tensor = torch.from_numpy(image_np).unsqueeze(0)

        return (image_tensor,)


#======图像层叠加
class ImageOverlayAlignment:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "image1": ("IMAGE", {"forceInput": True}),
                "image2": ("IMAGE", {"forceInput": True}),
                "alignment": (["top_left", "top_center", "top_right", "bottom_left", "bottom_center", "bottom_right", "center"], ),
                "scale": ("FLOAT", {"default": 1.0, "min": 0.1, "max": 10.0, "step": 0.1}),
                "opacity": ("FLOAT", {"default": 1.0, "min": 0.0, "max": 1.0, "step": 0.1}),
                "offset": ("STRING", {"default": "0,0,0,0", "multiline": False})
            }
        }

    RETURN_TYPES = ("IMAGE",)
    FUNCTION = "overlay_images"
    CATEGORY = "🇨🇳BOZO/PIC"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def overlay_images(self, image1, image2, alignment, offset, scale, opacity):
        image1_np = image1.cpu().numpy().squeeze()
        image2_np = image2.cpu().numpy().squeeze()
        img1 = Image.fromarray((image1_np * 255).astype(np.uint8)).convert('RGBA')
        img2 = Image.fromarray((image2_np * 255).astype(np.uint8)).convert('RGBA')
        img1_width, img1_height = img1.size
        new_width = int(img1_width * scale)
        new_height = int(img1_height * scale)
        img1 = img1.resize((new_width, new_height), Image.LANCZOS)
        img1 = self.adjust_opacity(img1, opacity)
        img1_width, img1_height = img1.size
        img2_width, img2_height = img2.size
        max_width = max(img1_width, img2_width)
        max_height = max(img1_height, img2_height)
        canvas = Image.new('RGBA', (max_width, max_height), (0, 0, 0, 0))
        if alignment == "top_left":
            img2_x, img2_y = 0, 0
        elif alignment == "top_center":
            img2_x = (max_width - img2_width) // 2
            img2_y = 0
        elif alignment == "top_right":
            img2_x = max_width - img2_width
            img2_y = 0
        elif alignment == "bottom_left":
            img2_x = 0
            img2_y = max_height - img2_height
        elif alignment == "bottom_center":
            img2_x = (max_width - img2_width) // 2
            img2_y = max_height - img2_height
        elif alignment == "bottom_right":
            img2_x = max_width - img2_width
            img2_y = max_height - img2_height
        elif alignment == "center":
            img2_x = (max_width - img2_width) // 2
            img2_y = (max_height - img2_height) // 2

        right_offset, left_offset, down_offset, up_offset = 0, 0, 0, 0
        offset_list = offset.split(',')
        if len(offset_list) >= 4:
            try:
                right_offset = int(offset_list[0]) if offset_list[0] else 0
                left_offset = int(offset_list[1]) if offset_list[1] else 0
                down_offset = int(offset_list[2]) if offset_list[2] else 0
                up_offset = int(offset_list[3]) if offset_list[3] else 0
            except ValueError:
                pass
        if alignment == "top_left":
            img1_x, img1_y = 0, 0
        elif alignment == "top_center":
            img1_x = (max_width - img1_width) // 2
            img1_y = 0
        elif alignment == "top_right":
            img1_x = max_width - img1_width
            img1_y = 0
        elif alignment == "bottom_left":
            img1_x = 0
            img1_y = max_height - img1_height
        elif alignment == "bottom_center":
            img1_x = (max_width - img1_width) // 2
            img1_y = max_height - img1_height
        elif alignment == "bottom_right":
            img1_x = max_width - img1_width
            img1_y = max_height - img1_height
        elif alignment == "center":
            img1_x = (max_width - img1_width) // 2
            img1_y = (max_height - img1_height) // 2
        img1_x += right_offset - left_offset
        img1_y += down_offset - up_offset
        img1_x = max(0, min(img1_x, max_width - img1_width))
        img1_y = max(0, min(img1_y, max_height - img1_height))
        img2_x = max(0, min(img2_x, max_width - img2_width))
        img2_y = max(0, min(img2_y, max_height - img2_height))
        canvas.paste(img2, (img2_x, img2_y), img2.split()[-1])
        canvas.paste(img1, (img1_x, img1_y), img1.split()[-1])
        output_image = np.array(canvas).astype(np.float32) / 255.0
        output_tensor = torch.from_numpy(output_image).unsqueeze(0)
        return (output_tensor,)

    def adjust_opacity(self, img, opacity):
        if opacity < 1.0:
            img = img.copy() 
            alpha = np.array(img.split()[-1]) * opacity
            alpha = alpha.astype(np.uint8)
            img.putalpha(Image.fromarray(alpha))
        return img


#======读取表格数据
class ReadExcelData:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "path/to/your/file.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_range": ("STRING", {"default": "2-3"}),
                "col_range": ("STRING", {"default": "1-4"}),
            },
             
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "read_excel_data"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def read_excel_data(self, excel_path, sheet_name, row_range, col_range):
        try:
            if "-" in row_range:
                start_row, end_row = map(int, row_range.split("-"))
            else:
                start_row = end_row = int(row_range)

            if "-" in col_range:
                start_col, end_col = map(int, col_range.split("-"))
            else:
                start_col = end_col = int(col_range)

            start_row = max(1, start_row)
            start_col = max(1, start_col)

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            output_lines = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                output_lines.append("|".join(row_data))

            workbook.close()
            del workbook

            return ("\n".join(output_lines),)

        except Exception as e:
            return (f"Error: {str(e)}",)


#======写入表格数据
class WriteExcelData:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "path/to/your/file.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}), 
                "row_range": ("STRING", {"default": "2-3"}),
                "col_range": ("STRING", {"default": "1-4"}),
                "data": ("STRING", {"default": "", "multiline": True}),
            },
             
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "write_excel_data"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def write_excel_data(self, excel_path, sheet_name, row_range, col_range, data):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}",)

            if not os.access(excel_path, os.W_OK):
                return (f"Error: No write permission for file at path: {excel_path}",)

            if "-" in row_range:
                start_row, end_row = map(int, row_range.split("-"))
            else:
                start_row = end_row = int(row_range)

            if "-" in col_range:
                start_col, end_col = map(int, col_range.split("-"))
            else:
                start_col = end_col = int(col_range)

            start_row = max(1, start_row)
            start_col = max(1, start_col)

            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            data_lines = data.strip().split("\n")

            for row_index, line in enumerate(data_lines, start=start_row):
                if row_index > end_row:
                    break

                cell_values = line.split("|")
                for col_index, cell_value in enumerate(cell_values, start=start_col):
                    if col_index > end_col:
                        break

                    if cell_value.strip():
                        sheet.cell(row=row_index, column=col_index).value = cell_value.strip()

            workbook.save(excel_path)

            workbook.close()
            del workbook

            return ("Data written successfully!",)

        except PermissionError as pe:
            return (f"Permission Error: {str(pe)}",)
        except Exception as e:
            return (f"Error: {str(e)}",)


#======图片插入表格
class WriteExcelImage:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "path/to/your/file.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_range": ("STRING", {"default": "1"}),
                "col_range": ("STRING", {"default": "1"}),
                "image_path": ("STRING", {"default": "path/to/your/image.png"}),
            },
             
        }

    RETURN_TYPES = ("STRING",)
    FUNCTION = "write_image"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def write_image(self, excel_path, sheet_name, row_range, col_range, image_path):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: Excel file does not exist at path: {excel_path}",)
            if not os.access(excel_path, os.W_OK):
                return (f"Error: No write permission for Excel file at path: {excel_path}",)
            if not os.path.exists(image_path):
                return (f"Error: Image file does not exist at path: {image_path}",)
            if not os.access(image_path, os.R_OK):
                return (f"Error: No read permission for image file at path: {image_path}",)
            if "-" in row_range:
                start_row, end_row = map(int, row_range.split("-"))
            else:
                start_row = end_row = int(row_range)
            if "-" in col_range:
                start_col, end_col = map(int, col_range.split("-"))
            else:
                start_col = end_col = int(col_range)
            start_row = max(1, start_row)
            start_col = max(1, start_col)
            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]
            thumbnail_size = (128, 128)
            with PILImage.open(image_path) as img:
                img = img.resize(thumbnail_size)
                img_byte_array = BytesIO()
                img.save(img_byte_array, format='PNG')
                img_byte_array.seek(0)
                openpyxl_img = OpenpyxlImage(img_byte_array)
            cell_address = openpyxl.utils.get_column_letter(start_col) + str(start_row)
            sheet.add_image(openpyxl_img, cell_address)
            workbook.save(excel_path)
            workbook.close()
            return ("Image inserted successfully!",)
        except PermissionError as pe:
            return (f"Permission Error: {str(pe)}",)
        except Exception as e:
            return (f"Error: {str(e)}",)


#======查找表格数据
class FindExcelData:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "path/to/your/file.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "search_content": ("STRING", {"default": "查找内容"}),
                "search_mode": (["精确查找", "模糊查找"], {"default": "精确查找"}),
            },
             
        }

    RETURN_TYPES = ("STRING", "INT", "INT")
    FUNCTION = "find_excel_data"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def find_excel_data(self, excel_path, sheet_name, search_content, search_mode):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}", None, None)
            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}", None, None)
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            results = []
            found_row = None
            found_col = None
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value if cell.value is not None else ""
                    cell_value_str = str(cell_value)
                    if (search_mode == "精确查找" and cell_value_str == search_content) or \
                       (search_mode == "模糊查找" and search_content in cell_value_str):
                        results.append(f"{sheet_name}|{row}|{col}|{cell_value}")
                        found_row = row
                        found_col = col

            workbook.close()
            del workbook
            if not results:
                return ("No results found.", None, None)
            return ("\n".join(results), found_row, found_col)
        except Exception as e:
            return (f"Error: {str(e)}", None, None)


#======读取表格数量差
class ReadExcelRowOrColumnDiff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "path/to/your/file.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "read_mode": (["读行", "读列"], {"default": "读行"}),
                "indices": ("STRING", {"default": "1,3"}),
            },
             
        }

    RETURN_TYPES = ("INT",)
    FUNCTION = "read_excel_row_or_column_diff"
    CATEGORY = "🇨🇳BOZO/文件"
    DESCRIPTION = ""
    def IS_CHANGED(*args, **kwargs): return float("NaN")

    def read_excel_row_or_column_diff(self, excel_path, sheet_name, read_mode, indices):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}",)

            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}",)

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(mode, index):
                count = 0
                if mode == "读行":
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=index, column=col).value
                        if cell_value is not None:
                            count += 1
                        else:
                            break
                elif mode == "读列":
                    for row in range(1, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row, column=index).value
                        if cell_value is not None:
                            count += 1
                        else:
                            break
                return count

            indices = indices.strip()
            if "," in indices:
                try:
                    index1, index2 = map(int, indices.split(","))
                except ValueError:
                    return (f"Error: Invalid indices format. Please use 'number,number' format.",)

                count1 = count_cells(read_mode, index1)
                count2 = count_cells(read_mode, index2)
                result = count1 - count2
            else:
                try:
                    index = int(indices)
                except ValueError:
                    return (f"Error: Invalid index format. Please enter a valid number.",)

                result = count_cells(read_mode, index)

            workbook.close()
            del workbook

            return (result,)

        except Exception as e:
            return (f"Error: {str(e)}",)


